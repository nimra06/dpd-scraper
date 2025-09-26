from __future__ import annotations
from typing import Optional, List, Dict, Tuple, Any

import os, time, re, json, string
from urllib.parse import urljoin
from dotenv import load_dotenv

import requests
from requests.exceptions import ReadTimeout, ConnectTimeout, Timeout, HTTPError
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from bs4 import BeautifulSoup
import pandas as pd

load_dotenv()

# ============================================================================
# CONFIG / CONSTANTS
# ============================================================================
BASE          = "https://health-products.canada.ca"
FORM_URL      = f"{BASE}/dpd-bdpp/?lang=eng"
RESULTS_URL   = f"{BASE}/dpd-bdpp/search-fast-recherche-rapide"
SEARCH_URL    = f"{BASE}/dpd-bdpp/search-recherche"
GET_PAGE_API  = f"{BASE}/dpd-bdpp/getNextPage"

TIMEOUT       = int(os.getenv("SCRAPER_TIMEOUT", "45"))
RETRIES       = int(os.getenv("SCRAPER_RETRIES", "3"))
RETRY_SLEEP   = float(os.getenv("SCRAPER_RETRY_SLEEP", "1.0"))

DEF_MAX_DEPTH          = int(os.getenv("SCRAPER_MAX_DEPTH", "1"))
DEF_TARGET_MIN_ROWS    = int(os.getenv("SCRAPER_TARGET_MIN_ROWS", "2000"))
DEF_ENRICH_FLUSH_EVERY = int(os.getenv("SCRAPER_ENRICH_FLUSH_EVERY", "50"))
DEF_REQUEST_SLEEP      = float(os.getenv("SCRAPER_REQUEST_SLEEP", "0.05"))
DEF_MAX_ROWS           = int(os.getenv("SCRAPER_MAX_ROWS", "0"))

SCRAPER_SWEEP_ORDER       = (os.getenv("SCRAPER_SWEEP_ORDER", "brand-first") or "brand-first").strip().lower()

# If SWEEP_PAGE_LIMIT <= 0, we drain until "stall" (no new rows) — no fixed limit required.
SWEEP_PAGE_LIMIT          = int(os.getenv("SCRAPER_SWEEP_PAGE_LIMIT", "0"))
PAGE_STALL_LIMIT          = int(os.getenv("SCRAPER_PAGE_STALL_LIMIT", "2"))
SCRAPER_SWEEP_MAX_EMPTY   = int(os.getenv("SCRAPER_SWEEP_MAX_EMPTY", "0"))
PAGE_KEYS = [k.strip() for k in os.getenv(
    "SCRAPER_PAGE_KEYS", "results_page,page,p,start,iDisplayStart"
).split(",") if k.strip()]

DEBUG                     = True
DEBUG_VERBOSE             = int(os.getenv("SCRAPER_DEBUG_VERBOSE", "1"))
LOG_EVERY_ADDED           = int(os.getenv("SCRAPER_LOG_EVERY_ADDED", "250"))
SWEEP_PREFIX_LOG_EVERY    = int(os.getenv("SCRAPER_SWEEP_PREFIX_LOG_EVERY", "1"))
DEBUG_PRINT_EVERY_ROW     = os.getenv("SCRAPER_DEBUG_PRINT_ROWS", "0") == "1"
DEBUG_PRINT_MAX_ROWS      = int(os.getenv("SCRAPER_DEBUG_PRINT_MAX", "200"))
DEBUG_PRINT_SAMPLE_EVERY  = int(os.getenv("SCRAPER_DEBUG_PRINT_SAMPLE_EVERY", "50"))

# --- Checkpointing (CSV) ---
SCRAPER_CHECKPOINT_EVERY_ROWS = int(os.getenv("SCRAPER_CHECKPOINT_EVERY_ROWS", "2000"))  # 0 = off
SCRAPER_CHECKPOINT_DIR        = os.getenv("SCRAPER_CHECKPOINT_DIR", "artifacts/checkpoints")
SCRAPER_CHECKPOINT_PREFIX     = os.getenv("SCRAPER_CHECKPOINT_PREFIX", "dpd")

_t0_global = time.time()
_last_beat = {"count": 0, "t": _t0_global}

def dbg(*args):
    if DEBUG:
        print("[DEBUG]", *args)

def dbg_row(stage: str, idx: int, row: dict):
    if not DEBUG_PRINT_EVERY_ROW:
        return
    if idx > DEBUG_PRINT_MAX_ROWS:
        return
    if idx > 20 and (idx % max(1, DEBUG_PRINT_SAMPLE_EVERY)) != 0:
        return
    filled = sum(1 for v in row.values() if v)
    din = row.get("DIN", "")
    brand = row.get("Product", "")
    comp = row.get("Company", "")
    print(f"[ROW {stage}] #{idx} DIN={din} brand={brand[:35]!r} company={comp[:35]!r} filled={filled}/{len(row)}")

# ============================================================================
# COLUMNS
# ============================================================================
DETAIL_COLS = [
    "Status","DIN URL","DIN","Company","Product","Class","PM See footnote1","Schedule",
    "# See footnote2","A.I. name See footnote3","Strength",
    "Current status date","Original market date",
    "Address","City","state","Country","Zipcode",
    "Number of active ingredient(s)","Biosimilar Biologic Drug",
    "American Hospital Formulary Service (AHFS)","Anatomical Therapeutic Chemical (ATC)",
    "Active ingredient group (AIG) number","Labelling","Product Monograph/Veterinary Date",
    "List of active ingredient","Dosage form","Route(s) of administration",
]
COLUMNS = DETAIL_COLS[:]

# ============================================================================
# UTILS
# ============================================================================
SPACES_RX = re.compile(r"[ \t\r\f\v]+")
DATE_RX   = re.compile(r"\b(19|20)\d{2}-\d{2}-\d{2}\b")

def norm(x: str | None) -> str:
    if not x:
        return ""
    x = x.replace("\xa0", " ").replace("\u200b", " ")
    return SPACES_RX.sub(" ", x).strip()

def _canon_din(d: str) -> str:
    return "".join(ch for ch in (d or "") if ch.isdigit())

def _canon_din_display(v: str) -> str:
    s = "".join(ch for ch in str(v or "") if ch.isdigit())
    return s or (v or "")

def _heartbeat(cum_rows: int, cap: int | None):
    if LOG_EVERY_ADDED <= 0:
        return
    if cum_rows >= _last_beat["count"] + LOG_EVERY_ADDED:
        now = time.time()
        dt = max(1e-6, now - _last_beat["t"])
        rate = (cum_rows - _last_beat["count"]) / dt
        elapsed = now - _t0_global
        msg = f"rows={cum_rows}"
        if cap:
            rem = max(0, cap - cum_rows)
            eta = rem / rate if rate > 0 else float("inf")
            msg += f" / cap={cap} | rate={rate:.1f}/s | elapsed={elapsed:.1f}s | eta~{eta:.1f}s"
        else:
            msg += f" | rate={rate:.1f}/s | elapsed={elapsed:.1f}s"
        dbg("[PROGRESS]", msg)
        _last_beat["count"] = cum_rows
        _last_beat["t"] = now

def _log_prefix_try(kind: str, val: str, ep_name: str, page: int, got: int, cum: int):
    if DEBUG_VERBOSE >= 2 or page == 1:
        dbg(f"[{kind}='{val}'] {ep_name} p{page}: {got} rows (cum={cum})")

def _coverage_counts(rows: List[Dict]) -> Dict[str, int]:
    counts = {k:0 for k in DETAIL_COLS}
    for r in rows:
        for k in counts:
            if r.get(k): counts[k] += 1
    return counts

def _ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)

def _write_checkpoint_csv(rows: List[Dict], cols: List[str], phase: str, count: int) -> str:
    """
    Writes a CSV checkpoint for 'rows'. Returns the file path.
    phase: 'list' or 'enriched'
    count: how many rows are in 'rows' (used in filename).
    """
    _ensure_dir(SCRAPER_CHECKPOINT_DIR)
    ts = time.strftime("%Y%m%d_%H%M%S")
    fn = f"{SCRAPER_CHECKPOINT_PREFIX}_{phase}_{count:06d}_{ts}.csv"
    path = os.path.join(SCRAPER_CHECKPOINT_DIR, fn)

    if cols:
        df = pd.DataFrame(rows, columns=cols)
    else:
        # union of keys (fallback)
        all_keys = set()
        for r in rows:
            all_keys.update(r.keys())
        ordered = [c for c in DETAIL_COLS if c in all_keys] + [k for k in sorted(all_keys) if k not in DETAIL_COLS]
        df = pd.DataFrame(rows, columns=ordered)

    df.to_csv(path, index=False, encoding="utf-8-sig")
    dbg(f"[CKPT] wrote {phase} CSV → {path}")
    return path

class _CheckpointGate:
    """Fires every SCRAPER_CHECKPOINT_EVERY_ROWS rows (if > 0)."""
    def __init__(self, every: int):
        self.every = max(0, int(every))
        self.last  = 0
    def maybe(self, rows_count: int) -> bool:
        if self.every <= 0:
            return False
        if rows_count >= self.last + self.every:
            self.last = rows_count
            return True
        return False

# ============================================================================
# HTTP / SESSION
# ============================================================================
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; dpd-scraper/1.6)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": FORM_URL,
        "Connection": "keep-alive",
    })
    retry = Retry(
        total=RETRIES,
        backoff_factor=RETRY_SLEEP,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET","POST"]),
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=16, pool_maxsize=64)
    s.mount("https://", adapter); s.mount("http://", adapter)
    return s

def _with_retries(fn):
    last = None
    for i in range(RETRIES):
        try:
            r = fn()
            r.raise_for_status()
            return r
        except (ReadTimeout, ConnectTimeout, Timeout, HTTPError, requests.exceptions.RequestException) as e:
            last = e
            time.sleep(RETRY_SLEEP * (i + 1))
    if last:
        raise last

def _fetch_page(sess: requests.Session, url: str, method: str, payload: dict,
                headers: dict | None = None) -> tuple[str, str]:
    method = (method or "GET").upper()
    if method == "POST":
        r = _with_retries(lambda: sess.post(url, data=payload, timeout=TIMEOUT, allow_redirects=True, headers=headers))
    else:
        r = _with_retries(lambda: sess.get(url, params=payload, timeout=TIMEOUT, allow_redirects=True, headers=headers))
    return r.text, method

# ============================================================================
# FORM BOOTSTRAP
# ============================================================================
def get_csrf(sess: requests.Session) -> str:
    r = _with_retries(lambda: sess.get(FORM_URL, timeout=TIMEOUT))
    soup = BeautifulSoup(r.text, "html.parser")
    el = soup.select_one("input#_csrf, input[name=_csrf]")
    return el["value"] if el and el.get("value") else ""

def _discover_form(sess: requests.Session) -> tuple[str, dict]:
    r = _with_retries(lambda: sess.get(FORM_URL, timeout=TIMEOUT, allow_redirects=True))
    soup = BeautifulSoup(r.text, "html.parser")

    def is_dpd_form(form) -> bool:
        names = {inp.get("name","") for inp in form.find_all("input")}
        expected = {"brandName","din","companyName","activeIngredient","aigNumber","biosimDrugSearch"}
        return any(k in names for k in expected)

    forms = soup.find_all("form")
    form = None
    for f in forms:
        if is_dpd_form(f):
            form = f; break
    if not form and forms:
        form = forms[0]
    if not form:
        return SEARCH_URL, {}

    action = form.get("action") or SEARCH_URL
    action_url = urljoin(BASE, action)
    if "canada.ca/en/sr/srb.html" in action_url:
        action_url = SEARCH_URL

    defaults = {}
    for inp in form.find_all("input"):
        name = inp.get("name")
        if not name: continue
        t = (inp.get("type") or "").lower()
        if t in ("checkbox","radio"):
            if inp.has_attr("checked"):
                defaults.setdefault(name, inp.get("value", "on"))
        else:
            defaults.setdefault(name, inp.get("value", ""))

    if "_csrf" not in defaults:
        try:
            csrf = get_csrf(sess)
            if csrf:
                defaults["_csrf"] = csrf
        except Exception:
            pass

    return action_url, defaults

def submit_search(sess: requests.Session, basic_html: bool = False) -> tuple[str, dict, str]:
    post_url, defaults = _discover_form(sess)
    base_filters = {
        "din": "", "atc": "", "companyName": "", "brandName": "",
        "activeIngredient": "", "aigNumber": "", "biosimDrugSearch": "",
    }
    payload = {**defaults, **base_filters}
    headers = {"Content-Type":"application/x-www-form-urlencoded","Origin":BASE,"Referer":FORM_URL}

    try:
        r = _with_retries(lambda: sess.post(post_url, data=payload, timeout=TIMEOUT, allow_redirects=True, headers=headers))
        if "canada.ca/en/sr/srb.html" in (r.url or ""):
            raise HTTPError("SRB relay after POST")
        html = r.text
    except Exception:
        r = _with_retries(lambda: sess.get(RESULTS_URL, params={"lang":"eng","wbdisable":"true"} if basic_html else {"lang":"eng"}, timeout=TIMEOUT))
        html = r.text

    return html, base_filters, SEARCH_URL

# ============================================================================
# LIST PAGE PARSER (HTML)
# ============================================================================
def parse_list_page_rows(html: str) -> list[dict]:
    soup = BeautifulSoup(html or "", "html.parser")
    table = soup.find("table", id="results")
    if not table: return []
    tbody = table.find("tbody")
    if not tbody: return []

    rows_out: list[dict] = []
    for tr in tbody.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 10:
            continue
        status    = tds[0].get_text(strip=True)
        din_cell  = tds[1]
        company   = tds[2].get_text(strip=True)
        product   = tds[3].get_text(strip=True)
        drugclass = tds[4].get_text(strip=True)
        pm        = tds[5].get_text(strip=True)
        schedule  = tds[6].get_text(strip=True)
        ai_num    = tds[7].get_text(strip=True)
        ai_name   = tds[8].get_text(strip=True)
        strength  = tds[9].get_text(strip=True)

        a_tag    = din_cell.find("a", href=True)
        din_text = (a_tag.get_text(strip=True) if a_tag else din_cell.get_text(strip=True))
        din_href = (a_tag["href"] if a_tag else "")
        din_url  = urljoin(BASE, din_href) if din_href else ""

        row = {
            "Status": status,
            "DIN URL": din_url,
            "DIN": _canon_din_display(din_text),
            "Company": company,
            "Product": product,
            "Class": drugclass,
            "PM See footnote1": pm,
            "Schedule": schedule,
            "# See footnote2": ai_num,
            "A.I. name See footnote3": ai_name,
            "Strength": strength,
        }
        rows_out.append(row)
    return rows_out

def _detect_table_paging(html: str) -> tuple[str, int]:
    try:
        soup = BeautifulSoup(html, "html.parser")
        tbl = soup.find("table", id="results")
        cfg = (tbl.get("data-wb-tables") or "") if tbl else ""
        src = re.search(r'"sAjaxSource"\s*:\s*"([^"]+)"', cfg)
        per = re.search(r'"iDisplayLength"\s*:\s*(\d+)', cfg)
        api = urljoin(BASE, src.group(1)) if src else GET_PAGE_API
        n   = int(per.group(1)) if per else 25
        return api, n
    except Exception:
        return GET_PAGE_API, 25

def _extract_total_entries(html: str) -> int | None:
    m = re.search(r"(?:of|sur)\s+([0-9][0-9\s,\.]*)\s+(?:entries|entrées)", html, flags=re.I)
    if not m: return None
    raw = m.group(1).replace("\u202f"," ").replace("\xa0"," ")
    raw = raw.replace(" ", "").replace(",", "").replace(".", "")
    try:
        return int(raw)
    except Exception:
        return None

# ============================================================================
# DT JSON → ROWS
# ============================================================================
def _dt_aa_to_listrows(aa: list) -> list[dict]:
    rows: list[dict] = []
    if not aa:
        return rows

    def _cell_text_and_href(cell_val) -> tuple[str, str]:
        s = "" if cell_val is None else str(cell_val)
        try:
            soup = BeautifulSoup(s, "html.parser")
            a = soup.find("a", href=True)
            href = a["href"] if a else ""
            text = soup.get_text(" ", strip=True)
            return text, href
        except Exception:
            text = re.sub(r"<[^>]+>", "", s)
            text = re.sub(r"\s+", " ", text).strip()
            return text, ""

    def _make_row(cells: list[str]) -> dict:
        def g(i): return cells[i] if i < len(cells) else ""
        din_txt, din_href = _cell_text_and_href(g(1))
        din_url = urljoin(BASE, din_href) if din_href else ""
        return {
            "Status": _cell_text_and_href(g(0))[0],
            "DIN URL": din_url,
            "DIN": _canon_din_display(din_txt),
            "Company": _cell_text_and_href(g(2))[0],
            "Product": _cell_text_and_href(g(3))[0],
            "Class": _cell_text_and_href(g(4))[0],
            "PM See footnote1": _cell_text_and_href(g(5))[0],
            "Schedule": _cell_text_and_href(g(6))[0],
            "# See footnote2": _cell_text_and_href(g(7))[0],
            "A.I. name See footnote3": _cell_text_and_href(g(8))[0],
            "Strength": _cell_text_and_href(g(9))[0],
        }

    if isinstance(aa[0], dict):
        DT_KEYS = ["status","din","company","brand","drugClass","pm","schedule","aiNum","majorAI","AIStrength"]
        for d in aa:
            cells = [d.get(k, "") for k in DT_KEYS]
            rows.append(_make_row(cells))
        return rows

    for row in aa:
        if not isinstance(row, (list, tuple)):
            continue
        rows.append(_make_row(list(row)))
    return rows

# ============================================================================
# ACTIVE INGREDIENTS (robust)
# ============================================================================
def _extract_ai_lines(soup: BeautifulSoup) -> list[str]:
    """
    Return a list like ["SODIUM CHLORIDE : 0.9 %", ...] from any of the
    known layouts (table caption, header, definition list, bullet list,
    or <br>-separated text). Handles English & French labels.
    """
    lines: list[str] = []

    def add(name: str, strength: str):
        name = norm(name)
        strength = norm(strength)
        if not name:
            return
        if name.lower() in ("name", "active ingredient", "ingrédient actif"):
            return
        lines.append(f"{name} : {strength}" if strength else name)

    # 1) Table with <caption> mentioning Active ingredient
    for cap in soup.find_all("caption"):
        cap_txt = norm(cap.get_text(" "))
        if "active ingredient" in cap_txt.lower() or "ingrédient actif" in cap_txt.lower():
            table = cap.find_parent("table")
            if table:
                for tr in table.find_all("tr"):
                    tds = tr.find_all(["td", "th"])
                    if len(tds) >= 2:
                        add(tds[0].get_text(" "), tds[1].get_text(" "))
            if lines:
                return lines

    # 2) Any table where header contains it
    for table in soup.find_all("table"):
        head_txt = norm((table.find("thead") or table).get_text(" "))
        if "active ingredient" in head_txt.lower() or "ingrédient actif" in head_txt.lower():
            for tr in table.find_all("tr"):
                tds = tr.find_all(["td", "th"])
                if len(tds) >= 2:
                    add(tds[0].get_text(" "), tds[1].get_text(" "))
            if lines:
                return lines

    # 3) Definition list pattern <dt>/<dd> or <th>/<td>
    for dt_tag in soup.find_all(["dt", "th"]):
        lab = norm(dt_tag.get_text(" "))
        if "active ingredient" in lab.lower() or "ingrédient actif" in lab.lower():
            dd = dt_tag.find_next("dd") or dt_tag.find_next("td")
            if dd:
                parts = [norm(li.get_text(" ")) for li in dd.find_all("li")]
                if not parts:
                    raw = dd.decode_contents()
                    parts = [norm(p) for p in re.split(r"<br\s*/?>|\n", raw, flags=re.I)]
                for p in parts:
                    if not p:
                        continue
                    if " : " in p:
                        nm, st = p.split(" : ", 1)
                        add(nm, st)
                    else:
                        m = re.match(r"^(.*?)[\s]{2,}(.*)$", p)
                        if m:
                            add(m.group(1), m.group(2))
                        else:
                            add(p, "")
            if lines:
                return lines

    # 4) Paragraph block that mentions it and contains <br>-separated items
    for p in soup.find_all("p"):
        text = p.get_text("\n", strip=True)
        if "active ingredient" in text.lower() or "ingrédient actif" in text.lower():
            parts = [norm(x) for x in text.split("\n") if norm(x)]
            for x in parts:
                if " : " in x:
                    nm, st = x.split(" : ", 1)
                    add(nm, st)
                else:
                    add(x, "")
            if lines:
                return lines

    return lines

def parse_active_ingredients_multiline(soup: BeautifulSoup) -> str:
    lines = _extract_ai_lines(soup)
    return "\n".join(lines)

# ============================================================================
# DETAIL PAGE (html.parser only)
# ============================================================================
def fetch_detail_fields(sess: requests.Session, din_url: str, sleep: float=0.0) -> Dict[str, str]:
    out = {k: "" for k in DETAIL_COLS}
    if not din_url:
        return out
    try:
        r = _with_retries(lambda: sess.get(din_url, timeout=TIMEOUT))
        soup = BeautifulSoup(r.text, "html.parser")

        def gr(label: str) -> str:
            lab = label.lower()
            for row in soup.select("div.row"):
                left = row.select_one("p.col-sm-4 strong")
                right = row.select_one("p.col-sm-8")
                if not left or not right: continue
                if lab in left.get_text(" ").lower():
                    return norm(right.get_text(" ").replace("\xa0"," "))
            return ""

        # core fields
        out["Status"] = gr("Status")
        out["Company"] = gr("Company")
        out["Product"] = gr("Product")
        out["Class"] = gr("Class")
        out["Schedule"] = gr("Schedule")
        out["Current status date"] = gr("Current status date")
        out["Original market date"] = gr("Original market date")
        out["Dosage form"] = gr("Dosage form")
        out["Route(s) of administration"] = gr("Route")
        out["Number of active ingredient(s)"] = gr("Number of active ingredient")
        out["American Hospital Formulary Service (AHFS)"] = gr("American Hospital Formulary Service")
        out["Anatomical Therapeutic Chemical (ATC)"] = gr("Anatomical Therapeutic Chemical")
        out["Active ingredient group (AIG) number"] = gr("Active ingredient group")

        # address block
        comp_row = None
        for row in soup.select("div.row"):
            left = row.select_one("p.col-sm-4 strong")
            if left and "company" in left.get_text(" ").lower():
                comp_row = row; break
        if comp_row:
            right = comp_row.select_one("p.col-sm-8")
            spans = [norm(s.get_text(" ")) for s in (right.select("span") if right else [])]
            if len(spans) >= 1: out["Address"] = spans[0]
            if len(spans) >= 2: out["City"]    = spans[1]
            if len(spans) >= 3: out["state"]   = spans[2]
            if len(spans) >= 4: out["Country"] = spans[3]
            if len(spans) >= 5: out["Zipcode"] = spans[4]

        # labelling (url + date)
        lab_url, lab_date = "", ""
        for row in soup.select("div.row"):
            left = row.select_one("p.col-sm-4 strong")
            right = row.select_one("p.col-sm-8")
            if not left or not right:
                continue
            label = left.get_text(" ", strip=True).lower()
            if any(key in label for key in ("product monograph/veterinary labelling","product monograph","labelling")):
                m = DATE_RX.search(right.get_text(" ", strip=True))
                if m: lab_date = m.group(0)
                a = right.find("a", href=True)
                if a and a["href"]:
                    href = a["href"]
                    lab_url = href if href.startswith("http") else urljoin(BASE, href)
                break
        out["Labelling"] = lab_url
        out["Product Monograph/Veterinary Date"] = lab_date

        # active ingredients (robust)
        ai_lines = _extract_ai_lines(soup)
        out["List of active ingredient"] = "\n".join(ai_lines)

        # If the AI list exists, make sure single AI fields are populated if blank
        if ai_lines:
            first = ai_lines[0]
            if " : " in first:
                nm, st = first.split(" : ", 1)
                if not out.get("A.I. name See footnote3"):
                    out["A.I. name See footnote3"] = nm.strip()
                if not out.get("Strength"):
                    out["Strength"] = st.strip()

        # biosimilar
        bs = gr("Biosimilar Biologic Drug")
        out["Biosimilar Biologic Drug"] = "Yes" if bs.lower().startswith("yes") else ("No" if bs else "")

        if sleep and sleep > 0:
            time.sleep(sleep)

        if out.get("List of active ingredient"):
            dbg(f"[AI] {din_url} -> {len(out['List of active ingredient'].splitlines())} AI line(s)")

        non_empty = sum(1 for v in out.values() if v)
        dbg(f"[DETAIL OK] filled {non_empty}/{len(out)} from {din_url}")
        return out
    except Exception as e:
        dbg(f"[DETAIL ERR] {din_url} :: {e!r}")
        return out

# ============================================================================
# SHARDING
# ============================================================================
BRAND_SYMBOL_PREFIXES = list("()[]{}#&+-,.'/")
def build_brand_prefixes() -> list[str]:
    return list(string.ascii_uppercase) + list(string.digits) + BRAND_SYMBOL_PREFIXES
def build_din_prefixes() -> list[str]:
    return list(string.digits)

# ============================================================================
# COLLECT (HTML p1, then DT JSON pages; drain fully per prefix)
# ============================================================================
def collect_all_list_rows(
    sess: requests.Session,
    first_html: str,
    base_filters: dict,
    endpoint_url: str,
    endpoint_method: str,
    min_rows: int,
    sleep: float,
    sample_n: Optional[int] = None,
    max_rows: int = 0,
    post_endpoint_url: Optional[str] = None,
) -> List[Dict]:
    rows_all: List[Dict] = []

    # p1
    page1_rows = parse_list_page_rows(first_html)
    rows_all.extend(page1_rows)
    dbg("Page 1 rows:", len(page1_rows))

    # coverage after p1
    if rows_all:
        c = _coverage_counts(rows_all)
        dbg("[LIST P1 COVERAGE]", c)

    page_api_url, per_page_dt = _detect_table_paging(first_html)
    per_page = per_page_dt if page1_rows else 25
    dbg(f"Detected paging API: {page_api_url} | per_page={per_page}")

    if sample_n:
        return rows_all[:sample_n]

    def _hit_cap() -> bool:
        return max_rows > 0 and len(rows_all) >= max_rows

    # CSV checkpoint gate for LIST phase
    _list_ckpt = _CheckpointGate(SCRAPER_CHECKPOINT_EVERY_ROWS)
    if _list_ckpt.maybe(len(rows_all)):
        _write_checkpoint_csv(rows_all, DETAIL_COLS, phase="list", count=len(rows_all))

    # helpers
    def _with_csrf(extra: dict) -> dict:
        token = ""
        try:
            token = get_csrf(sess)
        except Exception:
            token = base_filters.get("_csrf", "")
        data = {**base_filters, **extra}
        if token: data["_csrf"] = token
        data.setdefault("lang","eng"); data.setdefault("wbdisable","true")
        return data

    def _dt_fetch(start: int, filter_key: str, value: str) -> list[dict]:
        DT_COLS = ["status","din","company","brand","drugClass","pm","schedule","aiNum","majorAI","AIStrength"]
        if not hasattr(_dt_fetch, "_echo"): _dt_fetch._echo = 1
        echo = _dt_fetch._echo; _dt_fetch._echo += 1

        params = {
            "sEcho": str(echo),
            "iColumns": str(len(DT_COLS)),
            "sColumns": ",".join(DT_COLS),
            "iDisplayStart": str(start),
            "iDisplayLength": str(per_page),
            "iSortingCols": "1",
            "iSortCol_0": "3",
            "sSortDir_0": "asc",
            "sSearch": "", "bRegex": "false",
            "lang": "eng", "wbdisable": "true",
            "_": str(int(time.time()*1000)),
            filter_key: value,
        }
        for i, name in enumerate(DT_COLS):
            params[f"mDataProp_{i}"] = name
            params[f"sSearch_{i}"]   = ""
            params[f"bRegex_{i}"]    = "false"
            params[f"bSearchable_{i}"] = "true"
            params[f"bSortable_{i}"]   = "false" if name == "AIStrength" else "true"

        headers = {"Referer": FORM_URL, "X-Requested-With": "XMLHttpRequest"}
        dbg(f"[DT GET] {page_api_url} start={start} len={per_page} filter={filter_key}:{value}")
        r = _with_retries(lambda: sess.get(page_api_url, params=params, timeout=TIMEOUT, headers=headers))
        try:
            j = r.json()
        except Exception:
            j = json.loads(r.text)
        aa = j.get("aaData") or j.get("data") or []
        rows = _dt_aa_to_listrows(aa)
        if rows:
            dbg("[DT PAGE COVERAGE]", _coverage_counts(rows))
        return rows

    def _post_then_get(filter_key: str, value: str, page: int) -> tuple[list[dict], str]:
        if page == 1:
            payload = _with_csrf({filter_key: value})
            headers = {"Content-Type":"application/x-www-form-urlencoded","Origin":BASE,"Referer":FORM_URL}
            dbg(f"[POST try] {SEARCH_URL} {filter_key}='{value}'")
            rows = []
            try:
                r = _with_retries(lambda: sess.post(SEARCH_URL, data=payload, timeout=TIMEOUT, allow_redirects=True, headers=headers))
                if "canada.ca/en/sr/srb.html" in (r.url or ""):
                    dbg("[POST bounce] SRB relay")
                else:
                    rows = parse_list_page_rows(r.text)
                    dbg(f"[POST got] {len(rows)} rows")
            except Exception as e:
                dbg(f"[POST err] {e!r}")
            if not rows:
                html, _ = _fetch_page(sess, RESULTS_URL, "GET", {filter_key:value,"lang":"eng","wbdisable":"true"})
                rows = parse_list_page_rows(html)
                dbg(f"[GET p1 fallback] {len(rows)} rows")
            return rows, ("POST" if rows else "GET(HTML)")
        # DT pages
        start = (page - 1) * per_page
        try:
            rows = _dt_fetch(start, filter_key, value)
            if rows:
                return rows, "GET(JSON)"
        except Exception as e:
            dbg(f"[DT error] {e!r}")
        # HTML fallback
        base = {filter_key: value, "lang":"eng","wbdisable":"true"}
        for key in PAGE_KEYS:
            payload2 = dict(base)
            payload2[key] = page if key not in ("start","iDisplayStart") else start
            dbg(f"[HTML try] {RESULTS_URL} {key}={payload2[key]} filter={filter_key}:{value}")
            try:
                html, _ = _fetch_page(sess, RESULTS_URL, "GET", payload2)
                rows = parse_list_page_rows(html)
                if rows:
                    return rows, "GET(HTML)"
            except Exception as e:
                dbg(f"[HTML err] {e!r}")
        return [], "EMPTY"

    def _sweep_one(filter_key: str, value: str) -> Tuple[int, bool]:
        added_total = 0
        saw_any = False
        rows1, meth1 = _post_then_get(filter_key, value, 1)
        _log_prefix_try(filter_key, value, meth1, 1, len(rows1), len(rows_all))
        if not rows1:
            return 0, False
        saw_any = True
        seen = {_canon_din(r.get("DIN","")) for r in rows_all}
        added_this = 0
        for rr in rows1:
            din = _canon_din(rr.get("DIN",""))
            if din and din not in seen:
                rows_all.append(rr); seen.add(din)
                added_total += 1; added_this += 1
                dbg_row("LIST-P1", len(rows_all), rr)
                _heartbeat(len(rows_all), max_rows or None)
                if _hit_cap(): return added_total, True
        dbg(f"[PAGE] {filter_key}='{value}' p=1 got={len(rows1)} added_new={added_this} cum={len(rows_all)}")

        # Drain pages 2..∞
        stall = 0
        p = 2
        while True:
            rows_p, origin = _post_then_get(filter_key, value, p)
            _log_prefix_try(filter_key, value, origin, p, len(rows_p), len(rows_all))
            if not rows_p:
                stall += 1
                dbg(f"[PAGE] {filter_key}='{value}' p={p} got=0 (stall={stall}) cum={len(rows_all)}")
                if PAGE_STALL_LIMIT > 0 and stall >= PAGE_STALL_LIMIT:
                    dbg(f"[{filter_key}='{value}'] stopping after {stall} consecutive empty pages")
                    break
            else:
                stall = 0
                added_this = 0
                seen = {_canon_din(r.get("DIN","")) for r in rows_all}
                for rr in rows_p:
                    din = _canon_din(rr.get("DIN",""))
                    if din and din not in seen:
                        rows_all.append(rr); seen.add(din)
                        added_total += 1; added_this += 1
                        dbg_row("LIST-DT", len(rows_all), rr)
                        _heartbeat(len(rows_all), max_rows or None)
                        if _hit_cap(): return added_total, True
                dbg(f"[PAGE] {filter_key}='{value}' p={p} got={len(rows_p)} added_new={added_this} cum={len(rows_all)}")

            # checkpoint after each page
            if _list_ckpt.maybe(len(rows_all)):
                _write_checkpoint_csv(rows_all, DETAIL_COLS, phase="list", count=len(rows_all))

            # honor a fixed cap only if set (>0). SWEEP_PAGE_LIMIT<=0 means "no fixed page cap"
            if SWEEP_PAGE_LIMIT > 0 and p >= SWEEP_PAGE_LIMIT:
                dbg(f"[{filter_key}='{value}'] reached SWEEP_PAGE_LIMIT={SWEEP_PAGE_LIMIT} → break")
                break

            p += 1
            time.sleep(max(sleep, 0.02))
        return added_total, saw_any

    # sweep
    brand_prefixes = build_brand_prefixes()
    din_prefixes   = build_din_prefixes()

    def _run_prefix_group(kind: str, values: List[str]) -> bool:
        empty_streak = 0
        for i, v in enumerate(values, 1):
            if SWEEP_PREFIX_LOG_EVERY and i % SWEEP_PREFIX_LOG_EVERY == 0:
                dbg(f"[{kind}] prefix {i}/{len(values)} (cum={len(rows_all)})")
            added, saw_any = _sweep_one(kind, v)
            if not saw_any:
                empty_streak += 1
                if SCRAPER_SWEEP_MAX_EMPTY and empty_streak >= SCRAPER_SWEEP_MAX_EMPTY:
                    dbg(f"[{kind}] early stop after {empty_streak} empty prefixes")
                    return True
            else:
                empty_streak = 0
            if _hit_cap():
                return True
            time.sleep(max(sleep, 0.02))
        return False

    dbg("Starting sharded sweeps (POST→DT JSON→HTML)…")
    if SCRAPER_SWEEP_ORDER == "brand-first":
        if _run_prefix_group("brandName", brand_prefixes): return rows_all[:max_rows] if _hit_cap() else rows_all
        if _run_prefix_group("din", din_prefixes):         return rows_all[:max_rows] if _hit_cap() else rows_all
    else:
        if _run_prefix_group("din", din_prefixes):         return rows_all[:max_rows] if _hit_cap() else rows_all
        if _run_prefix_group("brandName", brand_prefixes): return rows_all[:max_rows] if _hit_cap() else rows_all

    return rows_all if not _hit_cap() else rows_all[:max_rows]

# ============================================================================
# ENRICHMENT (detail never clobbers non-empty)
# ============================================================================
def enrich_rows_with_details(sess: requests.Session, rows_all: list[dict], sleep: float) -> list[dict]:
    enriched: list[dict] = []
    _enrich_ckpt = _CheckpointGate(SCRAPER_CHECKPOINT_EVERY_ROWS)

    for i, r in enumerate(rows_all, 1):
        base = {k: "" for k in DETAIL_COLS}
        # copy list values
        for k, v in (r or {}).items():
            if v is not None:
                base[k] = str(v)

        din_url = (base.get("DIN URL") or "").strip()
        if din_url:
            try:
                det = fetch_detail_fields(sess, din_url, sleep=0.0)
            except Exception as e:
                dbg(f"[DETAIL] {din_url} error: {e!r}")
                det = {}
            for k, v in (det or {}).items():
                if v:  # only overwrite with non-empty
                    base[k] = v

        if not base.get("Biosimilar Biologic Drug"):
            base["Biosimilar Biologic Drug"] = "No"

        for col in DETAIL_COLS:
            base[col] = "" if base.get(col) is None else str(base.get(col))

        enriched.append(base)

        if i % DEF_ENRICH_FLUSH_EVERY == 0:
            filled = sum(1 for v in base.values() if v)
            dbg(f"[ENRICH] {i}/{len(rows_all)} (last row filled {filled}/{len(base)})")
        dbg_row("ENRICH", i, base)

        # checkpoint enriched set
        if _enrich_ckpt.maybe(len(enriched)):
            _write_checkpoint_csv(enriched, DETAIL_COLS, phase="enriched", count=len(enriched))

        if sleep and sleep > 0:
            time.sleep(sleep)

    return enriched

# ============================================================================
# EXCEL HELPER
# ============================================================================
def save_styled_excel(df: pd.DataFrame, xlsx_path: str) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="DPD")
        ws = xw.book["DPD"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 36
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for i, col in enumerate(df.columns, start=1):
            series = df[col].astype(str).fillna("")
            est = max(len(col), int(series.str.len().quantile(0.85)))
            est = max(12, min(est, 60))
            ws.column_dimensions[get_column_letter(i)].width = est

# ============================================================================
# ENTRYPOINT
# ============================================================================
def run_full_scrape(
    max_depth: int = DEF_MAX_DEPTH,
    target_min_rows: int = DEF_TARGET_MIN_ROWS,
    enrich_flush_every: int = DEF_ENRICH_FLUSH_EVERY,
    request_sleep: float = DEF_REQUEST_SLEEP,
    max_rows: int = DEF_MAX_ROWS,
) -> Tuple[List[Dict], Dict]:
    t0 = time.time()
    sess = make_session()

    first_html, base_filters, post_url = submit_search(sess, basic_html=True)

    list_rows = collect_all_list_rows(
        sess=sess,
        first_html=first_html,
        base_filters=base_filters,
        endpoint_url=RESULTS_URL,
        endpoint_method="GET",
        min_rows=target_min_rows,
        sleep=request_sleep,
        max_rows=max_rows,
        post_endpoint_url=post_url,
    )
    dbg("List collected:", len(list_rows))

    # coverage right after list phase (so you can catch empties early)
    dbg("[LIST COVERAGE TOTAL]", _coverage_counts(list_rows))

    enriched = enrich_rows_with_details(sess, list_rows, sleep=0.0)

    # coverage after enrichment
    dbg("[ENRICH COVERAGE TOTAL]", _coverage_counts(enriched))

    elapsed = round(time.time() - t0, 2)
    meta = {
        "strategy": "shard-sweeps(POST→DT→HTML) + full-detail-enrichment(html.parser)",
        "elapsed_sec": elapsed,
        "rows": len(enriched),
        "request_sleep": request_sleep,
        "max_depth": max_depth,
        "target_min_rows": target_min_rows,
        "max_rows": max_rows,
        "post_action": post_url,
        "columns_order": DETAIL_COLS,
    }
    dbg("Done. Rows:", len(enriched), "Elapsed(s):", elapsed)
    return enriched, meta

if __name__ == "__main__":
    rows, meta = run_full_scrape(
        target_min_rows=int(os.getenv("CLI_TARGET_MIN_ROWS", "30")),
        max_rows=int(os.getenv("CLI_MAX_ROWS", "30")),
        request_sleep=float(os.getenv("CLI_REQUEST_SLEEP", "0.03")),
    )
    print({"rows": len(rows), **meta})
