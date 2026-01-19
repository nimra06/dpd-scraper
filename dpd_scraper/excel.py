# dpd_scraper/excel.py
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

def save_styled_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DPD"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c in ws[r_idx]:
                c.font = Font(bold=True)
                c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r_idx].height = 28
        else:
            for c in ws[r_idx]:
                c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(10, length + 2), 60)

    wb.save(path)
