#!/usr/bin/env python3
"""
Convert an inspection XLSX export into CSV with sanitized column names.
Also writes a JSON mapping between original headers and sanitized names.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook


def sanitize_header(header: str, existing: Iterable[str]) -> str:
    base = re.sub(r"[^0-9a-zA-Z]+", "_", header).strip("_").lower()
    if not base:
        base = "col"
    candidate = base
    counter = 2
    existing_set = set(existing)
    while candidate in existing_set:
        candidate = f"{base}_{counter}"
        counter += 1
    return candidate


def convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path, mapping_path: Path) -> Dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError("Workbook contains no rows")

    mapping: "OrderedDict[str, str]" = OrderedDict()
    for header in headers:
        header_str = "" if header is None else str(header)
        sanitized = sanitize_header(header_str, mapping.values())
        mapping[header_str] = sanitized

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(mapping.values())
        for row in rows:
            normalized_row: List[str] = []
            for cell in row:
                if cell is None:
                    normalized_row.append("")
                else:
                    normalized_row.append(str(cell))
            # Ensure row length matches header length
            if len(normalized_row) < len(mapping):
                normalized_row.extend([""] * (len(mapping) - len(normalized_row)))
            elif len(normalized_row) > len(mapping):
                normalized_row = normalized_row[: len(mapping)]
            writer.writerow(normalized_row)

    mapping_path.parent.mkdir(parents=True, exist_ok=True)
    with mapping_path.open("w", encoding="utf-8") as mapping_file:
        json.dump(mapping, mapping_file, indent=2, ensure_ascii=False)

    return dict(mapping)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert inspection XLSX to CSV.")
    parser.add_argument("--xlsx", required=True, help="Path to source XLSX file.")
    parser.add_argument("--csv", required=True, help="Path to destination CSV file.")
    parser.add_argument(
        "--mapping",
        required=False,
        help="Path to JSON file storing header mapping (defaults to CSV path with .mapping.json).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)
    mapping_path = Path(args.mapping) if args.mapping else csv_path.with_suffix(".mapping.json")

    mapping = convert_xlsx_to_csv(xlsx_path, csv_path, mapping_path)
    print(f"Converted {xlsx_path} -> {csv_path}")
    print(f"Saved header mapping to {mapping_path}")
    print(f"Total columns: {len(mapping)}")


if __name__ == "__main__":
    main()

